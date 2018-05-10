# -*- coding: utf-8 -*-


def main():
    import sys
    import datetime
    import openpyxl
    import random
    import pandas as pd
    from dateutil.relativedelta import relativedelta

    DayOfTheWeek = ['月', 'Tue.', '水', '木', 'Fri.', '土', '日']
    Member = ["海東", "佐方", "菅原", "天野", "陳", "馮", "川崎", "加藤", "見内", "Yuan", "松本",
              "柴山", "鈴木", "薛", "沼田", "曽谷", "山本", "井上", "水間", "農見"]

    input_date = input('作成する年月度を入力してください（例：201608）：')
    try:
        # 当月の初日
        month_first = datetime.datetime.strptime(input_date, '%Y%m')
    except ValueError:
        input('年月を201608のように入力してください。')
        sys.exit()

    month_end = month_first + relativedelta(months=2)

    # 開始日と終了日
    date_start = month_first.replace(day=1)
    date_end = month_end - datetime.timedelta(days=1)

    # エクセルブックファイル（ひな形）
    file_template = 'cleaning_sheet_template.xlsx'
    book1 = openpyxl.load_workbook(file_template)
    # エクセルシート
    sheet1 = book1.get_sheet_by_name('Checklist')

    MemberOrder = []
    while len(MemberOrder) < 22*2:
        MemberOrder.extend(random.sample(Member, len(Member)))

    # 日の入力
    rowNum = 2
    memberNum = 0
    google_list = []
    while (rowNum < 22) and (date_start <= date_end):
        if date_start.weekday() == 1 or date_start.weekday() == 4:
            sheet1.cell(row=rowNum, column=1).value\
                = "{}/{}({})".format(date_start.month, date_start.day, DayOfTheWeek[date_start.weekday()])
            sheet1.cell(row=rowNum, column=2).value = MemberOrder[memberNum]
            sheet1.cell(row=rowNum, column=3).value = MemberOrder[memberNum+1]
            Subject = "cleaning duty[{}, {}]".format(MemberOrder[memberNum], MemberOrder[memberNum + 1])
            StartDate = "{:02}/{:02}/{}".format(date_start.month, date_start.day, date_start.year)
            AllDayEvent = "True"
            google_list.append([Subject, StartDate, AllDayEvent])
            rowNum += 1
            memberNum += 2

        date_start += datetime.timedelta(days=1)

    google = pd.DataFrame(google_list, columns=["Subject", "Start Date", "All Day Event"])
    # 新規ブックファイル保存
    file_new1 = month_first.strftime('%Y%m') + '_' + "cleaning_sheet.xlsx"
    book1.save(file_new1)
    google.to_csv(month_first.strftime("%Y%m") + "_" + "google_calender.csv", index=False, encoding="utf-8")


if __name__ == "__main__":
    main()
